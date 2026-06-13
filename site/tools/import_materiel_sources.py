#!/usr/bin/env python3
"""
Import EPI depuis MyDiveClub/_sourcesMatos/ vers Portail Club (matériel).

Dépendances : pip install openpyxl xlrd pyyaml

Usage :
  python site/tools/import_materiel_sources.py              # dry-run (défaut)
  python site/tools/import_materiel_sources.py --export payload.json
  python site/tools/import_materiel_sources.py --apply      # via PHP sur NAS/local (--update-states activé)
"""
from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import subprocess
import sys
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

try:
    import openpyxl
    import xlrd
    import yaml
except ImportError as exc:
    print("Dépendances manquantes : pip install openpyxl xlrd pyyaml", file=sys.stderr)
    raise SystemExit(1) from exc

ROOT = Path(__file__).resolve().parents[2]
SOURCES = ROOT / "_sourcesMatos"
MAPPING_FILE = SOURCES / "import_mapping.yaml"
APPLY_PHP = Path(__file__).resolve().parent / "import_materiel_apply.php"

SKIP_SHEETS = {"FR", "FORMULAIRE MAINTENANCE DETENDEUR"}
SKIP_SHEET_SUBSTR = ("liste menu", "feuil1")

GRADING_LEVELS = ["ras", "mineure", "majeure", "definitive"]

# Début de colonne RAS pour chaque critère (4 colonnes par critère)
GRADING_BLOCKS: dict[str, list[tuple[str, int]]] = {
    "bcd": [("sangles", 2), ("boucles", 6), ("confort", 10)],
    "mask": [("jupe", 2), ("sangle_boucles", 6), ("verres", 10)],
    "wetsuit": [("coutures", 2), ("fermeture", 6), ("revetement", 10)],
}

# Priorité d'état lors de la fusion (dossier Réparation vs Révision, etc.)
STATE_PRIORITY: dict[str, int] = {
    "operational": 1,
    "for_sale": 2,
    "in_repair": 3,
    "scrapped": 4,
}

# Segments de chemin → état courant (ordre : du plus proche du fichier au plus lointain)
FOLDER_STATE_RULES: list[tuple[re.Pattern[str], str]] = [
    (re.compile(r"rebut|scrap|definit", re.I), "scrapped"),
    (re.compile(r"repar", re.I), "in_repair"),
    (re.compile(r"vente", re.I), "for_sale"),
    (re.compile(r"revis", re.I), "operational"),
]


@dataclass
class Intervention:
    subtype: str
    done_on: str
    responsible_free: str
    summary: str = ""
    check_values: dict[str, str] = field(default_factory=dict)
    detail_json: dict[str, Any] | None = None

    def as_dict(self) -> dict[str, Any]:
        out: dict[str, Any] = {
            "subtype": self.subtype,
            "done_on": self.done_on,
            "responsible_free": self.responsible_free,
            "summary": self.summary,
        }
        if self.check_values:
            out["check_values"] = self.check_values
        if self.detail_json:
            out["detail_json"] = self.detail_json
        return out


@dataclass
class EquipmentItem:
    public_id: str
    structure_slug: str
    type_slug: str
    brand: str = ""
    model: str = ""
    serial: str = ""
    purchase_year: int | None = None
    state: str = "operational"
    state_from_folder: bool = False
    notes: str = ""
    interventions: list[Intervention] = field(default_factory=list)
    source: str = ""
    specs_json: dict[str, Any] | None = None

    def merge(self, other: "EquipmentItem") -> None:
        if other.brand and not self.brand:
            self.brand = other.brand
        if other.model and not self.model:
            self.model = other.model
        if other.serial and not self.serial:
            self.serial = other.serial
        if other.purchase_year and not self.purchase_year:
            self.purchase_year = other.purchase_year
        if other.notes and not self.notes:
            self.notes = other.notes
        if other.specs_json:
            if not self.specs_json:
                self.specs_json = other.specs_json
            else:
                merged_specs = dict(self.specs_json)
                for key, val in other.specs_json.items():
                    if val and not merged_specs.get(key):
                        merged_specs[key] = val
                self.specs_json = merged_specs
        self.state, self.state_from_folder = merge_equipment_states(
            self.state, self.state_from_folder, other.state, other.state_from_folder
        )
        self.interventions.extend(other.interventions)
        self.interventions.sort(key=lambda x: x.done_on)


def norm_folder_part(name: str) -> str:
    """Normalise un segment de chemin pour la détection d'état."""
    s = name.strip().lower()
    for src, dst in (
        ("é", "e"), ("è", "e"), ("ê", "e"), ("ë", "e"),
        ("à", "a"), ("â", "a"), ("ä", "a"),
        ("ù", "u"), ("û", "u"), ("ü", "u"),
        ("ô", "o"), ("ö", "o"),
        ("î", "i"), ("ï", "i"),
        ("ç", "c"),
    ):
        s = s.replace(src, dst)
    return s


def derive_state_from_path(path: Path) -> str | None:
    """
    Déduit l'état depuis un sous-dossier du chemin source.
    Ex. DET Révision → operational, DET Réparation → in_repair.
    """
    for part in reversed(path.parts):
        norm = norm_folder_part(part)
        if not norm or norm in {".", ".."}:
            continue
        for pattern, state in FOLDER_STATE_RULES:
            if pattern.search(norm):
                return state
    return None


def merge_equipment_states(
    state_a: str, folder_a: bool, state_b: str, folder_b: bool
) -> tuple[str, bool]:
    """Fusionne deux états — le nom de dossier prime sur le contenu Excel."""
    if folder_a and not folder_b:
        return state_a, True
    if folder_b and not folder_a:
        return state_b, True
    if folder_a and folder_b:
        pri_a = STATE_PRIORITY.get(state_a, 0)
        pri_b = STATE_PRIORITY.get(state_b, 0)
        if pri_a >= pri_b:
            return state_a, True
        return state_b, True
    pri_a = STATE_PRIORITY.get(state_a, 0)
    pri_b = STATE_PRIORITY.get(state_b, 0)
    return (state_a if pri_a >= pri_b else state_b), False


def apply_folder_state(item: EquipmentItem, path: Path) -> None:
    folder_state = derive_state_from_path(path)
    if folder_state is None:
        return
    item.state, item.state_from_folder = merge_equipment_states(
        item.state, item.state_from_folder, folder_state, True
    )


def norm_id(raw: Any) -> str:
    s = str(raw or "").strip().upper()
    s = re.sub(r"\s+", "", s)
    return s


def cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v == int(v):
        return str(int(v))
    return str(v).strip()


def parse_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, float):
        try:
            date_tuple = xlrd.xldate_as_tuple(v, 0)
            return dt.date(date_tuple[0], date_tuple[1], date_tuple[2]).isoformat()
        except Exception:
            return None
    s = str(v).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    m = re.match(r"^(\d{4}-\d{2}-\d{2})\b", s)
    if m:
        return m.group(1)
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", s)
    if m:
        d, mo, y = m.groups()
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    m = re.match(r"^(\d{1,2})\.(\d{1,2})\.(\d{2,4})$", s)
    if m:
        d, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
        return f"{y}-{int(mo):02d}-{int(d):02d}"
    m = re.match(r"^\.\./(\d{4})$", s)
    if m:
        return f"{m.group(1)}-01-01"
    return None


def year_from_date(v: Any) -> int | None:
    d = parse_date(v)
    return int(d[:4]) if d else None


def is_marked(row: list[Any], idx: int) -> bool:
    if idx >= len(row):
        return False
    v = row[idx]
    if v is None:
        return False
    s = str(v).strip().upper()
    return s in {"X", "O", "OUI", "OK", "1", "TRUE"} or s == "NEUF"


def cell_to_check_value(row: list[Any], idx: int, field_key: str) -> str | None:
    if idx >= len(row):
        return None
    v = row[idx]
    if v is None or str(v).strip() == "":
        return None
    if is_marked(row, idx):
        return "ok" if field_key == "ras" else "ok"
    s = str(v).strip().lower()
    if s in {"ko", "nok", "non"}:
        return "ko"
    if s in {"na", "n/a", "-"}:
        return "na"
    return "ok"


def check_profile_for_type(type_slug: str) -> str:
    if type_slug in {"bcd", "mask", "computer", "compass"}:
        return type_slug
    if type_slug.startswith("combi_") or type_slug.startswith("shorty_"):
        return "wetsuit"
    return ""


def grading_level_at(row: list[Any], start_col: int) -> str | None:
    for offset, level in enumerate(GRADING_LEVELS):
        if is_marked(row, start_col + offset):
            return level
    return None


def detect_grading_blocks(row: list[Any]) -> list[tuple[str, int]]:
    """Repère les blocs RAS/Mineure/Majeure/Définitive dans une ligne d'en-tête."""
    blocks: list[tuple[str, int]] = []
    i = 0
    while i < len(row):
        label = cell_str(row[i]).lower()
        if label in {"ras", "mineure", "majeure", "definitive", "définitive"}:
            blocks.append((f"criterion_{len(blocks)}", i))
            i += 4
            continue
        i += 1
    return blocks


def extract_grading_values(row: list[Any], type_slug: str) -> dict[str, str]:
    profile = check_profile_for_type(type_slug)
    if not profile or profile not in GRADING_BLOCKS:
        return {}
    out: dict[str, str] = {}
    for field_key, start_col in GRADING_BLOCKS[profile]:
        level = grading_level_at(row, start_col)
        if level:
            out[field_key] = level
    return out


# Legacy ordinateur (cases défaut + RAS global)
CHECK_COLUMNS_COMPUTER: list[tuple[int, str]] = [
    (2, "batterie_boutons"),
    (3, "couvercle"),
    (4, "ordi_hs"),
    (9, "ras"),
    (10, "mineure"),
    (11, "majeure"),
]


def extract_check_values(row: list[Any], type_slug: str) -> dict[str, str]:
    profile = check_profile_for_type(type_slug)
    if profile in GRADING_BLOCKS:
        return extract_grading_values(row, type_slug)
    if profile == "computer":
        out: dict[str, str] = {}
        for col_idx, field_key in CHECK_COLUMNS_COMPUTER:
            val = cell_to_check_value(row, col_idx, field_key)
            if val is not None:
                out[field_key] = val
        return out
    return {}


def derive_state_from_row(row: list[Any], check_values: dict[str, str] | None = None) -> str:
    checks = check_values or {}
    if any(v == "definitive" for v in checks.values()):
        return "scrapped"
    if any(v in {"mineure", "majeure"} for v in checks.values()):
        return "in_repair"
    text = " ".join(cell_str(c).upper() for c in row if c is not None)
    if "NEUF" in text:
        return "operational"
    if "DEFINITIVE" in text or "DÉFINITIVE" in text:
        return "scrapped"
    if "MAJEURE" in text or "MINEURE" in text:
        return "in_repair"
    return "operational"


def import_free_label(structure_slug: str) -> str:
    return "Import RÉDERIS" if structure_slug == "rederis" else "Import AQUABLUE"


def rows_from_openpyxl(ws) -> list[list[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=True)]


def rows_from_xlrd(sh) -> list[list[Any]]:
    out: list[list[Any]] = []
    for r in range(sh.nrows):
        out.append([sh.cell_value(r, c) for c in range(sh.ncols)])
    return out


def find_label_row(rows: list[list[Any]], label: str, col: int = 0) -> int | None:
    label_l = label.lower()
    for i, row in enumerate(rows):
        if col < len(row) and label_l in cell_str(row[col]).lower():
            return i
    return None


def find_intervention_columns(header_row: list[Any]) -> tuple[int, int]:
    tech_col = 0
    date_col: int | None = None
    for i, cell in enumerate(header_row):
        label = cell_str(cell).lower()
        if "date" in label and ("révision" in label or "revision" in label):
            date_col = i
            break
    if date_col is None:
        for i, cell in enumerate(header_row):
            if cell_str(cell).lower().startswith("date"):
                date_col = i
                break
    return tech_col, date_col if date_col is not None else 5


def should_skip_sheet(name: str) -> bool:
    upper = name.strip().upper()
    if upper in SKIP_SHEETS:
        return True
    low = name.strip().lower()
    return any(s in low for s in SKIP_SHEET_SUBSTR)


def is_valid_public_id(public_id: str) -> bool:
    return bool(public_id) and bool(re.match(r"^[A-Z0-9-]+$", public_id))


def sheet_fallback_id(sheet_name: str) -> str:
    cleaned = re.sub(r"^(GILET|GILETS|MASQUE|MASQUES|VETEMENT|SHORTY|ORDI)\s+", "", sheet_name.strip(), flags=re.I)
    cleaned = cleaned.strip()
    if cleaned:
        pid = norm_id(cleaned.split()[-1] if " " in cleaned else cleaned)
        if is_valid_public_id(pid):
            return pid
    pid = norm_id(sheet_name)
    return pid if is_valid_public_id(pid) else ""


def equipment_merge_key(item: EquipmentItem) -> str:
    return f"{item.structure_slug}:{item.type_slug}:{item.public_id}"


def parse_fiche_epi_rows(
    rows: list[list[Any]], sheet_name: str, structure_slug: str, type_slug: str, source: str
) -> EquipmentItem | None:
    title = cell_str(rows[2][0] if len(rows) > 2 and rows[2] else "")
    title_u = title.upper()
    if "FICHE DE GESTION" not in title_u and "FICHE" not in title_u:
        return None

    public_id = ""
    brand = ""
    model = ""
    serial = ""
    purchase_year = None

    for row in rows:
        c0 = cell_str(row[0] if len(row) > 0 else "")
        c1 = cell_str(row[1] if len(row) > 1 else "")
        c2 = cell_str(row[2] if len(row) > 2 else "")
        if c0.lower() in {"modèle", "modele"}:
            model = c1
            if "série" in c2.lower() or "serie" in c2.lower():
                serial = cell_str(row[3] if len(row) > 3 else "")
            elif c2.upper() in {"AQUALUNG", "SCUBAPRO", "MARES", "BEUCHAT", "CRESSI"}:
                brand = c2
        if c1.upper() == "AQUALUNG" and not brand:
            brand = "AQUALUNG"
        if c0.upper() in {"AQUABLUE PLONGEE", "REDERIS"} or c1.upper() in {"AQUABLUE PLONGEE", "REDERIS"}:
            ident = cell_str(row[2] if c0.upper() in {"AQUABLUE PLONGEE", "REDERIS"} else row[2])
            if ident and ident.lower() not in {"identification", ""}:
                public_id = norm_id(ident)
        if c0.lower().startswith("date achat"):
            purchase_year = year_from_date(row[1] if len(row) > 1 else None) or purchase_year
        if c0.lower() == "identification" and c1:
            public_id = norm_id(c1) or public_id
            if c2:
                brand = brand or c2

    if not public_id:
        public_id = sheet_fallback_id(sheet_name)

    if not is_valid_public_id(public_id):
        return None

    interventions: list[Intervention] = []
    state = "operational"
    header_idx = find_label_row(rows, "Nom du technicien")
    if header_idx is not None:
        tech_col, date_col = find_intervention_columns(rows[header_idx])
        for row in rows[header_idx + 1 : header_idx + 25]:
            tech = cell_str(row[tech_col] if len(row) > tech_col else "")
            if not tech or tech.lower().startswith("nom "):
                continue
            done = parse_date(row[date_col] if len(row) > date_col else None)
            if not done:
                continue
            check_values = extract_check_values(row, type_slug)
            summary_parts = []
            if check_values:
                summary_parts.append("Contrôle périodique importé")
            if "NEUF" in " ".join(cell_str(c) for c in row).upper():
                summary_parts.append("NEUF")
            summary = " — ".join(summary_parts) or "Révision importée"
            row_text = " ".join(cell_str(c) for c in row).upper()
            is_repair = bool(re.search(r"REPARATION|RÉPARATION", row_text))
            interventions.append(
                Intervention(
                    subtype="repair" if is_repair else "revision",
                    done_on=done,
                    responsible_free=tech,
                    summary=summary,
                    check_values=check_values if not is_repair else {},
                )
            )
            state = derive_state_from_row(row, check_values)

    return EquipmentItem(
        public_id=public_id,
        structure_slug=structure_slug,
        type_slug=type_slug,
        brand=brand,
        model=model,
        serial=serial,
        purchase_year=purchase_year,
        state=state,
        interventions=interventions,
        source=source,
    )


def parse_fiche_epi_workbook(wb_path: Path, structure_slug: str, type_slug: str) -> list[EquipmentItem]:
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    items: list[EquipmentItem] = []
    for name in wb.sheetnames:
        if should_skip_sheet(name):
            continue
        ws = wb[name]
        item = parse_fiche_epi_rows(rows_from_openpyxl(ws), name, structure_slug, type_slug, str(wb_path.name))
        if item:
            items.append(item)
    wb.close()
    return items


def parse_liste_epi_rows(rows: list[list[Any]], structure_slug: str, type_slug: str, source: str) -> list[EquipmentItem]:
    items: list[EquipmentItem] = []
    header_idx: int | None = None
    for i, row in enumerate(rows):
        c0 = cell_str(row[0] if row else "").lower()
        if c0 == "identification":
            header_idx = i
            break
    if header_idx is None:
        return items

    for row in rows[header_idx + 1 :]:
        ident = norm_id(row[0] if len(row) > 0 else "")
        if not ident:
            continue
        brand = cell_str(row[1] if len(row) > 1 else "")
        model = cell_str(row[2] if len(row) > 2 else "")
        done = parse_date(row[3] if len(row) > 3 else None)
        interventions: list[Intervention] = []
        if done:
            interventions.append(
                Intervention(
                    subtype="revision",
                    done_on=done,
                    responsible_free=import_free_label(structure_slug),
                    summary="Mise en service importée",
                    check_values={"controle_visuel": "ok", "fonctionnement": "ok"},
                )
            )
        items.append(
            EquipmentItem(
                public_id=ident,
                structure_slug=structure_slug,
                type_slug=type_slug,
                brand=brand,
                model=model,
                purchase_year=year_from_date(done),
                interventions=interventions,
                source=source,
            )
        )
    return items


def parse_liste_epi_workbook(wb_path: Path, structure_slug: str, type_slug: str) -> list[EquipmentItem]:
    wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
    items: list[EquipmentItem] = []
    for name in wb.sheetnames:
        items.extend(parse_liste_epi_rows(rows_from_openpyxl(wb[name]), structure_slug, type_slug, wb_path.name))
    wb.close()
    return items


def norm_cell_label(text: Any) -> str:
    return norm_folder_part(cell_str(text))


def col_val(row: list[Any], idx: int) -> str:
    if idx < 0 or idx >= len(row):
        return ""
    return cell_str(row[idx])


def find_label_col(row: list[Any], *fragments: str) -> int | None:
    for i, cell in enumerate(row):
        lab = norm_cell_label(cell)
        if not lab:
            continue
        if any(f in lab for f in fragments):
            return i
    return None


def value_after_label(row: list[Any], label_col: int) -> str:
    for off in (2, 1, 3, 4):
        v = col_val(row, label_col + off)
        if v and not _looks_like_form_header(v):
            return v
    return ""


def _looks_like_form_header(text: str) -> bool:
    lab = norm_cell_label(text)
    if not lab:
        return False
    headers = (
        "numero dossier", "numero de serie", "numero serie", "modele", "produit",
        "date", "nom du client", "configuration", "observations", "travail effectue",
        "cpn", "kit de maintenance", "valeur hp", "valeur mp", "effort", "numero de compte",
    )
    return any(h in lab for h in headers)


def fallback_public_id_from_source(source: str) -> str:
    stem = Path(source.split(":")[0]).stem
    pid = norm_id(stem)
    return pid if is_valid_public_id(pid) else ""


def _build_regulator_detail_json(
    tasks: dict[str, bool],
    observations: list[str],
    repairs: list[str],
    parts_changed: list[str],
    test_values: dict[str, Any],
    kits: dict[str, str],
) -> dict[str, Any]:
    tasks_other = [r for r in repairs if r and not any(tasks.values())]
    obs_text = " | ".join(observations + repairs)[:2000]
    return {
        "tasks": tasks,
        "tasks_other": tasks_other[:20],
        "observations": obs_text,
        "parts_changed": parts_changed,
        "test_values": test_values,
        "kits": kits,
    }


def parse_detendeur_rows(
    rows: list[list[Any]],
    sheet_name: str,
    structure_slug: str,
    type_slug: str,
    source: str,
    *,
    form_kind: str = "form",
) -> EquipmentItem | None:
    public_id = ""
    brand = "AQUALUNG"
    model = ""
    purchase_year: int | None = None
    observations: list[str] = []
    repairs: list[str] = []
    parts_changed: list[str] = []
    subtype = "revision" if form_kind == "form" else "repair"
    product_label = ""
    done_on: str | None = None
    technician = ""
    config_parts: list[str] = []
    in_config = False
    serial_parts: list[str] = []

    specs: dict[str, str] = {
        "model_hp": "",
        "model_mp": "",
        "model_octopus": "",
        "serial_hp": "",
        "serial_mp": "",
        "serial_octopus": "",
        "accessories": "",
        "product_label": "",
        "configuration": "",
    }
    tasks: dict[str, bool] = {
        "maint_hp": False,
        "maint_mp": False,
        "maint_octopus": False,
        "maint_gauge": False,
        "maint_hose": False,
        "direct_system": False,
        "hose_replaced": False,
        "nipple_replaced": False,
    }
    test_values: dict[str, Any] = {
        "hp_test": None,
        "mp_hp": None,
        "mp_open_effort": None,
        "mp_flow_effort": None,
        "oct_open_effort": None,
        "oct_flow_effort": None,
        "gauge_precision": None,
        "leak_test_ok": None,
    }
    kits: dict[str, str] = {"kit_hp_cpn": "", "kit_hp_lot": "", "kit_mp_cpn": "", "kit_mp_lot": ""}

    def mark_task(text: str) -> None:
        low = norm_cell_label(text)
        if not low:
            return
        if ("1er" in low and "etage" in low) or "maintenance 1er" in low:
            tasks["maint_hp"] = True
        if ("2" in low and "etage" in low) or "maintenance 2" in low:
            tasks["maint_mp"] = True
        if "octopus" in low:
            tasks["maint_octopus"] = True
        if "manometre" in low:
            tasks["maint_gauge"] = True
        if "flexible" in low and "remplac" not in low:
            tasks["maint_hose"] = True
        if "direct system" in low:
            tasks["direct_system"] = True
        if "flexible remplac" in low:
            tasks["hose_replaced"] = True
        if "nipple" in low:
            tasks["nipple_replaced"] = True

    def append_unique(bucket: list[str], text: str) -> None:
        t = text.strip()
        if t and t not in bucket:
            bucket.append(t)

    for ri, row in enumerate(rows):
        dossier_col = find_label_col(row, "numero dossier")
        if dossier_col is not None:
            public_id = norm_id(value_after_label(row, dossier_col)) or public_id

        date_col = find_label_col(row, "date")
        if date_col is not None:
            lab = norm_cell_label(row[date_col])
            if lab == "date" or (lab.startswith("date") and "achat" not in lab and "technicien" not in lab):
                if date_col <= 2:
                    purchase_year = year_from_date(value_after_label(row, date_col)) or purchase_year

        produit_col = find_label_col(row, "produit")
        if produit_col is not None and "detendeur" not in norm_cell_label(row[produit_col]):
            val = value_after_label(row, produit_col)
            if val:
                product_label = val
                model = val

        model1_col = find_label_col(row, "modele 1", "1er etage")
        model2_col = find_label_col(row, "modele 2", "2eme etage", "2eme etage")
        is_model_header = False
        if model1_col is not None and model2_col is not None:
            lab1 = norm_cell_label(row[model1_col])
            lab2 = norm_cell_label(row[model2_col])
            is_model_header = "modele" in lab1 and "modele" in lab2
        if is_model_header and ri + 1 < len(rows):
            vr = rows[ri + 1]
            specs["model_hp"] = col_val(vr, 1) or specs["model_hp"]
            specs["model_mp"] = col_val(vr, 3) or specs["model_mp"]
            specs["model_octopus"] = col_val(vr, 4) or col_val(vr, 5) or specs["model_octopus"]
            specs["accessories"] = col_val(vr, 6) or col_val(vr, 7) or specs["accessories"]

        serial_labels = sum(
            1
            for cell in row
            if "numero" in norm_cell_label(cell) and "serie" in norm_cell_label(cell)
        )
        if serial_labels >= 2 and ri + 1 < len(rows):
            vr = rows[ri + 1]
            specs["serial_hp"] = col_val(vr, 1) or specs["serial_hp"]
            specs["serial_mp"] = col_val(vr, 3) or specs["serial_mp"]
            specs["serial_octopus"] = col_val(vr, 4) or col_val(vr, 5) or specs["serial_octopus"]
            for s in (specs["serial_hp"], specs["serial_mp"], specs["serial_octopus"]):
                if s and s not in serial_parts:
                    serial_parts.append(s)

        if find_label_col(row, "configuration detendeur") is not None:
            in_config = True
            continue
        if in_config:
            if find_label_col(row, "reparation") is not None or find_label_col(row, "travail effectue") is not None:
                in_config = False
            else:
                for idx in (1, 4):
                    part = col_val(row, idx)
                    if part and "observations" not in norm_cell_label(part):
                        append_unique(config_parts, part)

        maint_date_col = find_label_col(row, "le (date)", "le date")
        if maint_date_col is not None and maint_date_col + 1 < len(row):
            parsed_maint = parse_date(row[maint_date_col + 1])
            if parsed_maint:
                done_on = parsed_maint

        tech_col = find_label_col(row, "technicien", "par (nom")
        if tech_col is not None:
            technician = value_after_label(row, tech_col) or technician

        for ci, cell in enumerate(row):
            raw = cell_str(cell)
            lab = norm_cell_label(cell)
            if not raw:
                continue

            if lab == "reparation":
                rep = value_after_label(row, ci)
                if rep:
                    append_unique(repairs, rep)

            if ci == 2 and "maintenance" in lab:
                append_unique(repairs, raw)
                mark_task(raw)

            if ci == 4 and raw and not lab.startswith("autre"):
                mark_task(raw)
                if any(k in lab for k in ("remplac", "joint", "valve", "schrader", "cpn")):
                    append_unique(repairs, raw)

            if raw.upper().startswith("CPN") or (lab.startswith("cpn") and ":" in raw):
                append_unique(parts_changed, raw)

            if "observations du technicien" in lab:
                obs = value_after_label(row, ci)
                if obs:
                    append_unique(observations, obs)
            elif ci >= 6 and "observations" not in lab and len(raw) > 20 and not _looks_like_form_header(raw):
                if in_config or find_label_col(row, "configuration detendeur") is not None:
                    append_unique(observations, raw)

            if "valeur hp" in lab:
                test_values["hp_test"] = _parse_float(value_after_label(row, ci))
            if "valeur mp" in lab and "1er" in lab:
                test_values["mp_hp"] = _parse_float(value_after_label(row, ci))
            if "effort" in lab and "ouverture" in lab and "2" in lab:
                test_values["mp_open_effort"] = _parse_float(value_after_label(row, ci))
            if "effort" in lab and "flux" in lab and "2" in lab:
                test_values["mp_flow_effort"] = _parse_float(value_after_label(row, ci))
            if "effort" in lab and "octopus" in lab and "ouverture" in lab:
                test_values["oct_open_effort"] = _parse_float(value_after_label(row, ci))
            if "effort" in lab and "octopus" in lab and "flux" in lab:
                test_values["oct_flow_effort"] = _parse_float(value_after_label(row, ci))
            if "precision manometre" in lab:
                test_values["gauge_precision"] = _parse_float(value_after_label(row, ci))
            if "etancheite" in lab:
                test_values["leak_test_ok"] = is_marked(row, ci + 2) or bool(value_after_label(row, ci))

            if "kit 1" in lab and "etage" in lab:
                kits["kit_hp_lot"] = col_val(row, ci + 2) or kits["kit_hp_lot"]
                kits["kit_hp_cpn"] = col_val(row, ci + 1) or kits["kit_hp_cpn"]
            if "kit 2" in lab and "etage" in lab:
                kits["kit_mp_lot"] = col_val(row, ci + 2) or kits["kit_mp_lot"]
                kits["kit_mp_cpn"] = col_val(row, ci + 1) or kits["kit_mp_cpn"]

            mark_task(raw)

    if not public_id:
        public_id = norm_id(sheet_name) or fallback_public_id_from_source(source)

    if not public_id or public_id in SKIP_SHEETS:
        return None

    specs["product_label"] = product_label or model
    specs["configuration"] = ", ".join(config_parts)
    specs["public_id"] = public_id

    if not done_on:
        done_on = f"{purchase_year or 2020}-01-01"

    summary = " | ".join([s for s in observations + repairs if s]) or (
        "Maintenance détendeur importée" if form_kind == "form" else "Réparation détendeur importée"
    )
    if form_kind == "repair" or re.search(r"repar", summary, re.I):
        subtype = "repair"
    elif form_kind == "form":
        subtype = "revision"

    detail_json = _build_regulator_detail_json(
        tasks, observations, repairs, parts_changed, test_values, kits
    )

    responsible = technician.strip() or import_free_label(structure_slug)

    return EquipmentItem(
        public_id=public_id,
        structure_slug=structure_slug,
        type_slug=type_slug,
        brand=brand,
        model=model,
        serial=" / ".join(serial_parts) if serial_parts else " / ".join(
            s for s in (specs["serial_hp"], specs["serial_mp"], specs["serial_octopus"]) if s
        ),
        purchase_year=purchase_year,
        state="in_repair" if form_kind == "repair" else "operational",
        specs_json=specs,
        interventions=[
            Intervention(
                subtype=subtype,
                done_on=done_on,
                responsible_free=responsible,
                summary=summary[:2000],
                detail_json=detail_json,
            )
        ],
        source=source,
    )


def _parse_float(raw: Any) -> float | None:
    s = cell_str(raw).replace(",", ".")
    if not s:
        return None
    try:
        return float(re.sub(r"[^\d.\-]", "", s) or "0")
    except ValueError:
        return None


def parse_detendeur_xls(path: Path, structure_slug: str, type_slug: str) -> list[EquipmentItem]:
    wb = xlrd.open_workbook(str(path))
    items: list[EquipmentItem] = []
    for name in wb.sheet_names():
        if should_skip_sheet(name):
            continue
        sh = wb.sheet_by_name(name)
        item = parse_detendeur_rows(
            rows_from_xlrd(sh), name, structure_slug, type_slug, f"{path.name}:{name}", form_kind="form"
        )
        if item:
            items.append(item)
    return items


def parse_detendeur_xlsx(path: Path, structure_slug: str, type_slug: str, *, form_kind: str = "repair") -> list[EquipmentItem]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    items: list[EquipmentItem] = []
    for name in wb.sheetnames:
        if should_skip_sheet(name):
            continue
        item = parse_detendeur_rows(
            rows_from_openpyxl(wb[name]), name, structure_slug, type_slug, path.name, form_kind=form_kind
        )
        if item:
            items.append(item)
    wb.close()
    return items


def load_mapping() -> dict[str, Any]:
    if not MAPPING_FILE.is_file():
        raise FileNotFoundError(f"Mapping introuvable : {MAPPING_FILE}")
    return yaml.safe_load(MAPPING_FILE.read_text(encoding="utf-8"))


def collect_items() -> dict[str, EquipmentItem]:
    mapping = load_mapping()
    merged: dict[str, EquipmentItem] = {}

    def add_item(item: EquipmentItem) -> None:
        key = equipment_merge_key(item)
        if key in merged:
            merged[key].merge(item)
        else:
            merged[key] = item

    for folder_name, cfg in mapping.items():
        folder = SOURCES / folder_name
        if not folder.is_dir():
            print(f"AVERTissement : dossier absent {folder}")
            continue
        structure_slug = cfg["structure_slug"]
        type_slug = cfg["type_slug"]
        parser = cfg.get("parser")

        if parser == "fiche_epi":
            file_name = cfg["file"]
            matches = list(folder.glob(file_name))
            for fp in matches:
                if fp.suffix.lower() == ".pdf":
                    continue
                for item in parse_fiche_epi_workbook(fp, structure_slug, type_slug):
                    apply_folder_state(item, fp)
                    add_item(item)
        elif parser == "liste_epi":
            file_name = cfg["file"]
            for fp in folder.glob(file_name):
                for item in parse_liste_epi_workbook(fp, structure_slug, type_slug):
                    apply_folder_state(item, fp)
                    add_item(item)
        elif cfg.get("sources"):
            for src in cfg["sources"]:
                rel = src["path"]
                src_parser = src["parser"]
                for fp in folder.glob(rel):
                    if src_parser == "detendeur_form" and fp.suffix.lower() == ".xls":
                        items = parse_detendeur_xls(fp, structure_slug, type_slug)
                    elif src_parser == "detendeur_form":
                        items = parse_detendeur_xlsx(fp, structure_slug, type_slug, form_kind="form")
                    elif src_parser == "detendeur_repair":
                        items = parse_detendeur_xlsx(fp, structure_slug, type_slug, form_kind="repair")
                    else:
                        continue
                    for item in items:
                        apply_folder_state(item, fp)
                        add_item(item)

    return merged


def payload_from_items(items: dict[str, EquipmentItem]) -> dict[str, Any]:
    return {
        "items": [
            {
                "public_id": it.public_id,
                "structure_slug": it.structure_slug,
                "type_slug": it.type_slug,
                "brand": it.brand,
                "model": it.model,
                "serial": it.serial,
                "purchase_year": it.purchase_year,
                "state": it.state,
                "notes": it.notes,
                "specs_json": it.specs_json,
                "interventions": [i.as_dict() for i in it.interventions],
                "source": it.source,
            }
            for it in sorted(items.values(), key=lambda x: (x.structure_slug, x.type_slug, x.public_id))
        ]
    }


def payload_sync_regulators_from_items(items: dict[str, EquipmentItem]) -> dict[str, Any]:
    """Payload backfill détendeurs existants — specs, meta, état, interventions."""
    regulators = [it for it in items.values() if it.type_slug == "regulator"]
    return {
        "sync_regulators": True,
        "items": [
            {
                "public_id": it.public_id,
                "structure_slug": it.structure_slug,
                "type_slug": it.type_slug,
                "brand": it.brand,
                "model": it.model,
                "serial": it.serial,
                "purchase_year": it.purchase_year,
                "state": it.state,
                "specs_json": it.specs_json,
                "interventions": [i.as_dict() for i in it.interventions],
                "source": it.source,
            }
            for it in sorted(regulators, key=lambda x: (x.structure_slug, x.public_id))
        ],
    }


def print_sync_regulators_report(items: dict[str, EquipmentItem]) -> None:
    regulators = [it for it in items.values() if it.type_slug == "regulator"]
    if not regulators:
        print("Sync détendeurs : aucun détendeur parsé.")
        return
    with_specs = sum(
        1 for it in regulators
        if it.specs_json and any(v for k, v in it.specs_json.items() if k != "public_id" and v)
    )
    with_detail = sum(
        1 for it in regulators
        if any(inv.detail_json for inv in it.interventions)
    )
    print(f"\nSync détendeurs — {len(regulators)} équipement(s) parsé(s) :")
    print(f"  specs_json renseigné : {with_specs}")
    print(f"  detail_json intervention : {with_detail}")
    print(f"  interventions totales : {sum(len(it.interventions) for it in regulators)}")
    print("\nÉchantillon (10 premiers) :")
    for it in sorted(regulators, key=lambda x: x.public_id)[:10]:
        s = it.specs_json or {}
        hint = s.get("product_label") or s.get("model_hp") or "—"
        print(f"  {it.public_id} [{it.structure_slug}] {hint} — {len(it.interventions)} int. — {it.state}")


def payload_sync_states_from_items(items: dict[str, EquipmentItem]) -> dict[str, Any]:
    """Payload minimal : uniquement items dont l'état vient d'un sous-dossier source."""
    folder_items = [it for it in items.values() if it.state_from_folder]
    return {
        "sync_states": True,
        "items": [
            {
                "public_id": it.public_id,
                "structure_slug": it.structure_slug,
                "type_slug": it.type_slug,
                "state": it.state,
                "source": it.source,
            }
            for it in sorted(folder_items, key=lambda x: (x.structure_slug, x.type_slug, x.public_id))
        ],
    }


STATE_LABELS = {
    "operational": "Opérationnel",
    "in_repair": "En réparation",
    "scrapped": "Rebut",
    "for_sale": "À vendre",
}


def print_sync_states_report(items: dict[str, EquipmentItem]) -> None:
    folder_items = [it for it in items.values() if it.state_from_folder]
    if not folder_items:
        print("Sync états : aucun item avec état déduit d'un sous-dossier (Révision/Réparation, etc.).")
        return
    by_state: dict[str, int] = {}
    for it in folder_items:
        by_state[it.state] = by_state.get(it.state, 0) + 1
    print(f"\nSync états (dossier) — {len(folder_items)} équipement(s) :")
    for state, cnt in sorted(by_state.items(), key=lambda x: -x[1]):
        label = STATE_LABELS.get(state, state)
        print(f"  {label} ({state}) : {cnt}")
    print("\nDétail (état cible depuis _sourcesMatos) :")
    for it in sorted(folder_items, key=lambda x: (x.structure_slug, x.type_slug, x.public_id)):
        label = STATE_LABELS.get(it.state, it.state)
        src = f" — {it.source}" if it.source else ""
        print(f"  {it.public_id} [{it.structure_slug}/{it.type_slug}] -> {label}{src}")


def normalize_tech_name(name: str) -> str:
    """Approximation aliases PHP pour rapport dry-run."""
    aliases = {
        "courtaudiere": "Julien C.",
        "coutaudiere": "Julien C.",
        "courdaudiere": "Julien C.",
        "courtaidiere": "Julien C.",
        "cputaudiere": "Julien C.",
        "mesnier": "Marie M.",
        "petit": "Eric D.",
        "delmas": "Eric D.",
        "enjalbert": "Eric D.",
        "vidal": "Eric D.",
        "roynard": "Eric D.",
        "le provost": "Eric D.",
        "leprovost": "Eric D.",
        "jeremie georges": "Eric D.",
        "lacote": "Julie L.",
        "chaumeton": "Jérôme C.",
    }
    key = name.strip().lower()
    return aliases.get(key, name.strip())


def print_report(items: dict[str, EquipmentItem]) -> None:
    by_type: dict[str, int] = {}
    by_structure: dict[str, int] = {}
    by_state: dict[str, int] = {}
    int_count = 0
    checks_count = 0
    unknown_techs: set[str] = set()
    known = {
        "julien c.", "marie m.", "eric d.", "julie l.", "nicolas c.", "jérôme c.", "jerome c.",
        "import aquablue", "import réderis", "import rederis",
    }

    for it in items.values():
        by_type[it.type_slug] = by_type.get(it.type_slug, 0) + 1
        by_structure[it.structure_slug] = by_structure.get(it.structure_slug, 0) + 1
        by_state[it.state] = by_state.get(it.state, 0) + 1
        for inv in it.interventions:
            int_count += 1
            if inv.check_values:
                checks_count += 1
            tech = normalize_tech_name(inv.responsible_free).lower()
            if tech and tech not in known and not tech.startswith("import "):
                unknown_techs.add(inv.responsible_free.strip())

    print(f"Équipements : {len(items)}")
    for slug, cnt in sorted(by_structure.items()):
        print(f"  [{slug}] {cnt}")
    for slug, cnt in sorted(by_type.items()):
        print(f"  - {slug}: {cnt}")
    for slug, cnt in sorted(by_state.items()):
        print(f"  état {slug}: {cnt}")
    print(f"Interventions : {int_count} ({checks_count} avec critères)")

    if unknown_techs:
        print(f"\nTechniciens non aliasés ({len(unknown_techs)}) :")
        for t in sorted(unknown_techs)[:20]:
            print(f"  - {t}")
        if len(unknown_techs) > 20:
            print(f"  … et {len(unknown_techs) - 20} autres")

    print("\nÉchantillon (10 premiers IDs) :")
    for key in sorted(items.keys())[:10]:
        it = items[key]
        folder_hint = " [dossier]" if it.state_from_folder else ""
        print(f"  {it.public_id} ({it.structure_slug}/{it.type_slug}) — {it.state}{folder_hint} — {len(it.interventions)} int.")


def run_apply(
    payload: dict[str, Any],
    *,
    update_states: bool = True,
    sync_states_only: bool = False,
    sync_regulators_only: bool = False,
) -> int:
    if not APPLY_PHP.is_file():
        print(f"Script PHP introuvable : {APPLY_PHP}", file=sys.stderr)
        return 1
    cmd = ["php", str(APPLY_PHP)]
    if sync_regulators_only:
        cmd.append("--sync-regulators-only")
    elif sync_states_only:
        cmd.append("--sync-states-only")
    elif update_states:
        cmd.append("--update-states")
    proc = subprocess.run(
        cmd,
        input=json.dumps(payload, ensure_ascii=False),
        text=True,
        encoding="utf-8",
        capture_output=True,
        cwd=str(APPLY_PHP.parent),
    )
    if proc.stdout:
        print(proc.stdout)
    if proc.stderr:
        print(proc.stderr, file=sys.stderr)
    return proc.returncode


def main() -> int:
    parser = argparse.ArgumentParser(description="Import EPI _sourcesMatos → Portail Club")
    parser.add_argument("--apply", action="store_true", help="Appliquer via import_materiel_apply.php")
    parser.add_argument("--export", metavar="FILE", help="Exporter le payload JSON")
    parser.add_argument(
        "--sync-states",
        action="store_true",
        help="Sync états uniquement (sous-dossiers Révision/Réparation) — pas de création ni interventions",
    )
    parser.add_argument(
        "--sync-regulators",
        action="store_true",
        help="Backfill détendeurs existants (specs_json, meta, état, interventions) — Excel fait foi",
    )
    args = parser.parse_args()

    if not SOURCES.is_dir():
        print(f"Dossier sources introuvable : {SOURCES}", file=sys.stderr)
        return 1

    items = collect_items()
    if not items:
        print("Aucun équipement parsé.", file=sys.stderr)
        return 1

    if args.sync_regulators:
        print_sync_regulators_report(items)
        payload = payload_sync_regulators_from_items(items)
        if not payload["items"]:
            print("\nRien à synchroniser.", file=sys.stderr)
            return 1
        if args.export:
            out = Path(args.export)
            out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"\nPayload sync détendeurs exporté : {out}")
        if args.apply:
            print("\nApplication sync détendeurs en base…")
            return run_apply(payload, sync_regulators_only=True)
        if not args.export:
            print("\nMode dry-run sync détendeurs — aucune écriture. Utilisez --apply ou --export.")
        return 0

    if args.sync_states:
        print_sync_states_report(items)
        payload = payload_sync_states_from_items(items)
        if not payload["items"]:
            print("\nRien à synchroniser.", file=sys.stderr)
            return 1
        if args.export:
            out = Path(args.export)
            out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
            print(f"\nPayload sync exporté : {out}")
        if args.apply:
            print("\nApplication sync états en base…")
            return run_apply(payload, sync_states_only=True)
        if not args.export:
            print("\nMode dry-run sync — aucune écriture. Utilisez --apply ou --export.")
        return 0

    print_report(items)
    payload = payload_from_items(items)

    if args.export:
        out = Path(args.export)
        out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"\nPayload exporté : {out}")

    if args.apply:
        print("\nApplication en base…")
        return run_apply(payload)

    if not args.export:
        print("\nMode dry-run — aucune écriture. Utilisez --apply ou --export.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
